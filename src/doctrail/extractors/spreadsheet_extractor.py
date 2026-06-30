"""Extraction helpers for Microsoft Excel spreadsheets."""

from __future__ import annotations

import logging
import os
import shutil
import subprocess
import tempfile
import time as time_mod
import csv
import fcntl
from datetime import date, datetime, time
from io import StringIO
from pathlib import Path
from typing import Iterable

from bs4 import BeautifulSoup
import chardet

logger = logging.getLogger(__name__)
SOFFICE_SLOT_DIR = Path("/tmp/doctrail-soffice-slots")


def _format_scalar(value) -> str:
    """Convert spreadsheet cell values into stable plain text."""
    if value is None:
        return ""

    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"

    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")

    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, time):
        return value.isoformat(timespec="seconds")

    if isinstance(value, float) and value.is_integer():
        return str(int(value))

    return str(value).strip()


def _format_row(values: Iterable[object]) -> str:
    """Render one spreadsheet row with trailing empty cells removed."""
    rendered = [_format_scalar(value) for value in values]

    while rendered and not rendered[-1]:
        rendered.pop()

    if not any(rendered):
        return ""

    return " | ".join(rendered)


def _render_sheet(sheet_name: str, rows: Iterable[Iterable[object]]) -> str:
    """Render a worksheet into plain text."""
    lines = [line for line in (_format_row(row) for row in rows) if line]
    if not lines:
        return ""

    return f"Sheet: {sheet_name}\n" + "\n".join(lines)


def _normalize_rows(rows: Iterable[Iterable[object]]) -> list[list[str]]:
    return [[_format_scalar(value) for value in row] for row in rows]


def _rows_to_csv(rows: list[list[str]]) -> str:
    buffer = StringIO()
    writer = csv.writer(buffer)
    for row in rows:
        writer.writerow(row)
    return buffer.getvalue()


def _sheet_payload(sheet_name: str, rows: Iterable[Iterable[object]]) -> dict:
    normalized_rows = _normalize_rows(rows)
    return {
        "sheet_name": sheet_name,
        "rows": normalized_rows,
        "csv": _rows_to_csv(normalized_rows),
        "text": _render_sheet(sheet_name, normalized_rows),
    }


class _SofficeSlot:
    def __init__(self, fd: int):
        self.fd = fd

    def close(self) -> None:
        try:
            fcntl.flock(self.fd, fcntl.LOCK_UN)
        finally:
            os.close(self.fd)


def _acquire_soffice_slot(max_procs: int, timeout: float = 300.0) -> _SofficeSlot:
    SOFFICE_SLOT_DIR.mkdir(parents=True, exist_ok=True)
    deadline = time_mod.time() + timeout

    while time_mod.time() < deadline:
        for slot_index in range(max_procs):
            slot_path = SOFFICE_SLOT_DIR / f"slot-{slot_index}.lock"
            fd = os.open(slot_path, os.O_CREAT | os.O_RDWR, 0o600)
            try:
                fcntl.flock(fd, fcntl.LOCK_EX | fcntl.LOCK_NB)
                return _SofficeSlot(fd)
            except BlockingIOError:
                os.close(fd)
                continue
        time_mod.sleep(0.1)

    raise TimeoutError(f"Timed out waiting for a LibreOffice slot after {timeout:.0f}s")


def _resolve_soffice() -> str | None:
    """Return the LibreOffice CLI path if available."""
    return shutil.which("soffice") or shutil.which("libreoffice")


def _read_prefix(file_path: str, size: int = 4096) -> bytes:
    with open(file_path, "rb") as handle:
        return handle.read(size)


def _read_text_with_detected_encoding(file_path: str) -> str:
    raw_bytes = Path(file_path).read_bytes()
    if not raw_bytes:
        return ""

    sample = raw_bytes[:65536]
    detected = chardet.detect(sample).get("encoding")
    candidate_encodings = [
        detected,
        "utf-8-sig",
        "utf-16",
        "utf-16le",
        "utf-16be",
        "gb18030",
        "gbk",
        "big5",
        "latin-1",
    ]

    for encoding in candidate_encodings:
        if not encoding:
            continue
        try:
            return raw_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue

    return raw_bytes.decode("utf-8", errors="ignore")


def _looks_like_html_spreadsheet(file_path: str) -> bool:
    """Detect HTML files saved with a spreadsheet extension."""
    prefix = _read_prefix(file_path).lstrip().lower()
    return (
        prefix.startswith(b"<!doctype html")
        or prefix.startswith(b"<html")
        or prefix.startswith(b"<!--")
        or prefix.startswith(b"<script")
        or b"<html" in prefix
        or b"<table" in prefix
    )


def _extract_structured_from_html_spreadsheet(file_path: str) -> list[dict]:
    """Extract structured table data from HTML-backed spreadsheets."""
    html = Path(file_path).read_text(encoding="utf-8", errors="ignore")
    soup = BeautifulSoup(html, "html.parser")
    title = soup.title.get_text(" ", strip=True) if soup.title else Path(file_path).stem

    rendered_tables = []
    for index, table in enumerate(soup.find_all("table"), start=1):
        rows = []
        for tr in table.find_all("tr"):
            cells = [cell.get_text(" ", strip=True) for cell in tr.find_all(["th", "td"])]
            if any(cells):
                rows.append(cells)

        if rows:
            sheet_name = title if index == 1 else f"{title} ({index})"
            rendered_tables.append(_sheet_payload(sheet_name, rows))

    if rendered_tables:
        return rendered_tables

    fallback_text = soup.get_text("\n", strip=True).strip()
    if fallback_text:
        return [{
            "sheet_name": title,
            "rows": [[fallback_text]],
            "csv": _rows_to_csv([[fallback_text]]),
            "text": f"Sheet: {title}\n{fallback_text}",
        }]

    return []


def extract_structured_from_delimited(file_path: str) -> tuple[str, str, list[dict]]:
    """
    Extract text from CSV/TSV-style files.

    Returns:
        (content, extraction_method, structured_sheets)
    """
    if _looks_like_html_spreadsheet(file_path):
        sheets = _extract_structured_from_html_spreadsheet(file_path)
        if sheets:
            return "\n\n".join(sheet["text"] for sheet in sheets).strip(), "html_spreadsheet", sheets

    text = _read_text_with_detected_encoding(file_path)
    if not text.strip():
        return "", "csv", []

    default_delimiter = "\t" if Path(file_path).suffix.lower() == ".tsv" else ","
    sample = "\n".join(text.splitlines()[:20]).strip()

    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = default_delimiter

    reader = csv.reader(StringIO(text), delimiter=delimiter)
    rows = [row for row in reader if any(str(cell).strip() for cell in row)]
    if not rows:
        return "", "csv", []

    payload = _sheet_payload(Path(file_path).stem, rows)
    payload["delimiter"] = delimiter
    payload["source_format"] = Path(file_path).suffix.lower().lstrip(".")
    return payload["text"], "csv", [payload]


def _convert_with_soffice(file_path: str, target_ext: str) -> str:
    """Convert an Office document using LibreOffice and return the converted path."""
    soffice = _resolve_soffice()
    if not soffice:
        raise RuntimeError(
            f"LibreOffice is not installed. Install `soffice` to convert {Path(file_path).suffix} files."
        )

    max_procs = int(os.environ.get("DOCTRAIL_SOFFICE_MAX_PROCS", "2"))
    retry_count = max(0, int(os.environ.get("DOCTRAIL_SOFFICE_RETRIES", "1")))
    slot = _acquire_soffice_slot(max(1, max_procs))
    try:
        with tempfile.TemporaryDirectory(prefix="doctrail-office-") as tmpdir:
            last_error: RuntimeError | None = None
            for attempt in range(retry_count + 1):
                profile_dir = Path(tmpdir) / f"lo-profile-{attempt}"
                profile_dir.mkdir(exist_ok=True)
                cmd = [
                    soffice,
                    "--headless",
                    "--nologo",
                    "--nodefault",
                    "--nolockcheck",
                    f"-env:UserInstallation={profile_dir.as_uri()}",
                    "--convert-to",
                    target_ext,
                    "--outdir",
                    tmpdir,
                    file_path,
                ]
                logger.debug("Running soffice: %s", " ".join(cmd))

                try:
                    result = subprocess.run(cmd, capture_output=True, text=True, timeout=180)
                except subprocess.TimeoutExpired as exc:
                    if attempt < retry_count:
                        logger.warning(
                            "LibreOffice timed out converting %s on attempt %d/%d; retrying",
                            Path(file_path).name,
                            attempt + 1,
                            retry_count + 1,
                        )
                        time_mod.sleep(0.5 * (attempt + 1))
                        continue
                    raise RuntimeError(
                        f"LibreOffice timed out converting {Path(file_path).name} to {target_ext}"
                    ) from exc

                converted_path = Path(tmpdir) / f"{Path(file_path).stem}.{target_ext}"
                if not converted_path.exists():
                    matches = sorted(Path(tmpdir).glob(f"*.{target_ext}"))
                    if matches:
                        converted_path = matches[0]

                if not converted_path.exists():
                    stderr = result.stderr.strip() or result.stdout.strip()
                    last_error = RuntimeError(
                        f"LibreOffice failed to convert {Path(file_path).name} to {target_ext}: "
                        f"returncode={result.returncode} {stderr}".strip()
                    )
                    should_retry = attempt < retry_count and (result.returncode < 0 or not stderr)
                    if should_retry:
                        logger.warning(
                            "LibreOffice produced no output for %s on attempt %d/%d (returncode=%s); retrying",
                            Path(file_path).name,
                            attempt + 1,
                            retry_count + 1,
                            result.returncode,
                        )
                        time_mod.sleep(0.5 * (attempt + 1))
                        continue
                    raise last_error

                if result.returncode != 0:
                    logger.warning(
                        "LibreOffice returned %s for %s but produced %s; continuing with converted file",
                        result.returncode,
                        Path(file_path).name,
                        converted_path.name,
                    )

                persistent_tmpdir = tempfile.mkdtemp(prefix="doctrail-office-converted-")
                persistent_path = Path(persistent_tmpdir) / converted_path.name
                shutil.copy2(converted_path, persistent_path)
                return str(persistent_path)

            if last_error is not None:
                raise last_error
            raise RuntimeError(f"LibreOffice failed to convert {Path(file_path).name} to {target_ext}")
    finally:
        slot.close()


def _extract_structured_from_xlsx_openpyxl(file_path: str, *, data_only: bool) -> list[dict]:
    """Read workbook text and structured rows using openpyxl."""
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("openpyxl is not installed. Run `uv sync` to ingest .xlsx files.") from exc

    workbook = openpyxl.load_workbook(
        filename=file_path,
        read_only=True,
        data_only=data_only,
    )
    try:
        rendered_sheets = []
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            payload = _sheet_payload(sheet.title, rows)
            if payload["text"]:
                rendered_sheets.append(payload)
        return rendered_sheets
    finally:
        workbook.close()


def extract_structured_from_xlsx(file_path: str) -> tuple[str, str, list[dict]]:
    """
    Extract text from an .xlsx workbook.

    Returns:
        (content, extraction_method, structured_sheets)
    """
    if _looks_like_html_spreadsheet(file_path):
        sheets = _extract_structured_from_html_spreadsheet(file_path)
        if sheets:
            return "\n\n".join(sheet["text"] for sheet in sheets).strip(), "html_spreadsheet", sheets

    sheets = _extract_structured_from_xlsx_openpyxl(file_path, data_only=True)
    content = "\n\n".join(sheet["text"] for sheet in sheets).strip()
    if content:
        return content, "openpyxl", sheets

    # Formula-only workbooks may not have cached values.
    sheets = _extract_structured_from_xlsx_openpyxl(file_path, data_only=False)
    content = "\n\n".join(sheet["text"] for sheet in sheets).strip()
    if content:
        return content, "openpyxl_formula", sheets

    return "", "openpyxl", []


def extract_text_from_xlsx(file_path: str) -> tuple[str, str]:
    content, method, _ = extract_structured_from_xlsx(file_path)
    return content, method


def _extract_structured_from_xls_xlrd(file_path: str) -> list[dict]:
    """Read legacy .xls workbooks using xlrd."""
    try:
        import xlrd
    except ImportError as exc:
        raise RuntimeError("xlrd is not installed. Run `uv sync` to ingest .xls files.") from exc

    workbook = xlrd.open_workbook(file_path, on_demand=True)
    try:
        rendered_sheets = []
        for sheet in workbook.sheets():
            normalized_rows = []
            for row_index in range(sheet.nrows):
                row_values = []
                for column_index in range(sheet.ncols):
                    cell = sheet.cell(row_index, column_index)
                    value = cell.value
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        value = xlrd.xldate.xldate_as_datetime(value, workbook.datemode)
                    elif cell.ctype == xlrd.XL_CELL_BOOLEAN:
                        value = bool(value)
                    row_values.append(value)

                if any(value not in (None, "") for value in row_values):
                    normalized_rows.append(row_values)

            if not normalized_rows:
                continue

            payload = _sheet_payload(sheet.name, normalized_rows)
            if payload["text"]:
                rendered_sheets.append(payload)
        return rendered_sheets
    finally:
        workbook.release_resources()


def extract_structured_from_xls(file_path: str) -> tuple[str, str, list[dict]]:
    """
    Extract text from a legacy .xls workbook.

    Returns:
        (content, extraction_method, structured_sheets)
    """
    if _looks_like_html_spreadsheet(file_path):
        sheets = _extract_structured_from_html_spreadsheet(file_path)
        if sheets:
            return "\n\n".join(sheet["text"] for sheet in sheets).strip(), "html_spreadsheet", sheets

    xlrd_error = None
    try:
        sheets = _extract_structured_from_xls_xlrd(file_path)
        content = "\n\n".join(sheet["text"] for sheet in sheets).strip()
        if content:
            return content, "xlrd", sheets
    except Exception as exc:
        xlrd_error = exc

    converted_path = None
    soffice_error = None
    try:
        converted_path = _convert_with_soffice(file_path, "xlsx")
        content, _, sheets = extract_structured_from_xlsx(converted_path)
        if content:
            return content, "soffice_to_xlsx_openpyxl", sheets
    except Exception as exc:
        soffice_error = exc
    finally:
        if converted_path:
            converted = Path(converted_path)
            try:
                converted.unlink(missing_ok=True)
                converted.parent.rmdir()
            except OSError:
                pass

    error_parts = []
    if xlrd_error is not None:
        error_parts.append(f"xlrd failed: {xlrd_error}")
    if soffice_error is not None:
        error_parts.append(f"LibreOffice conversion failed: {soffice_error}")

    if error_parts:
        raise RuntimeError("; ".join(error_parts))

    return "", "xlrd", []


def extract_text_from_xls(file_path: str) -> tuple[str, str]:
    content, method, _ = extract_structured_from_xls(file_path)
    return content, method
